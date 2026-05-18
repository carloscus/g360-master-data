#!/usr/bin/env python3
"""
G360 Plantilla Precios Reader v1.0
Lee plantilla_precios.xlsx (3 hojas: PRODUCTOS, DESCUENTOS, PRECIOS_FIJOS)
Genera JSON compatible con el flujo de trabajo G360
Autor: CCUSI | v1.0.0
"""
import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: pip install openpyxl"); sys.exit(1)

def leer_plantilla_precios(ruta):
    """Lee el formato plantilla_precios.xlsx: 3+ hojas (PRODUCTOS, DESCUENTOS, PRECIOS_FIJOS, SKU_CLIENTES opcional)"""
    precios, errs = [], []
    sku_clientes_data = {}
    
    try:
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        
        # Verificar hojas requeridas
        hojas_requeridas = ['PRODUCTOS', 'DESCUENTOS', 'PRECIOS_FIJOS']
        for hoja in hojas_requeridas:
            if hoja not in wb.sheetnames:
                errs.append(f"Falta hoja '{hoja}'")
        
        if errs:
            return precios, errs
        
        # =====================================================
        # Leer hoja SKU_CLIENTES (opcional)
        # =====================================================
        if 'SKU_CLIENTES' in wb.sheetnames:
            ws_sku_clientes = wb['SKU_CLIENTES']
            headers_sku = [str(c.value).strip().lower() if c.value else "" for c in ws_sku_clientes[1]]
            idx_sku = {h: i for i, h in enumerate(headers_sku) if h}
            
            if 'sku' in idx_sku and 'sku_cliente' in idx_sku:
                for row in ws_sku_clientes.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    sku = row[idx_sku.get('sku', 0)]
                    sku_cliente = row[idx_sku.get('sku_cliente', 1)]
                    if sku and sku_cliente:
                        sku_clientes_data[str(sku).strip()] = str(sku_cliente).strip()
                
                print(f"Códigos de cliente leídos: {len(sku_clientes_data)}")
            else:
                print("Advertencia: Hoja SKU_CLIENTES sin columnas 'sku' y 'sku_cliente'")
        
        # =====================================================
        # Leer hoja PRODUCTOS (para obtener orden y SKU)
        # =====================================================
        ws_productos = wb['PRODUCTOS']
        productos_data = {}
        
        # Obtener encabezados
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws_productos[1]]
        
        # Verificar columnas requeridas
        columnas_requeridas = ['orden', 'sku']
        for col in columnas_requeridas:
            if col not in headers:
                errs.append(f"Falta columna '{col}' en hoja PRODUCTOS")
        
        if errs:
            return precios, errs
        
        # Crear mapeo de índices
        idx = {h: i for i, h in enumerate(headers) if h}
        
        # Leer productos
        for row in ws_productos.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            
            orden = row[idx.get('orden', 0)]
            sku = row[idx.get('sku', 1)]
            
            if sku and orden:
                productos_data[str(sku).strip()] = int(orden) if orden else 0
        
        print(f"Productos leídos: {len(productos_data)}")
        
        # =====================================================
        # Leer hoja DESCUENTOS (solo los que están en catálogo)
        # =====================================================
        ws_descuentos = wb['DESCUENTOS']
        descuentos_data = {}
        
        # Obtener encabezados
        headers_desc = [str(c.value).strip().lower() if c.value else "" for c in ws_descuentos[1]]
        idx_desc = {h: i for i, h in enumerate(headers_desc) if h}
        
        # Verificar columna SKU
        if 'sku' not in idx_desc:
            errs.append("Falta columna 'sku' en hoja DESCUENTOS")
            return precios, errs
        
        # Leer descuentos (solo los que están en catálogo)
        descuentos_filtrados = 0
        descuentos_ignorados = 0
        
        for row in ws_descuentos.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            
            sku = row[idx_desc.get('sku', 0)]
            if not sku:
                continue
            
            sku_str = str(sku).strip()
            
            # FILTRO: Solo incluir descuentos de SKUs que están en PRODUCTOS
            if sku_str not in productos_data:
                descuentos_ignorados += 1
                continue
            
            descuentos_filtrados += 1
            
            # Obtener descuentos (convertir de porcentaje a decimal)
            desc1 = row[idx_desc.get('desc1', 1)] if 'desc1' in idx_desc else 0
            desc2 = row[idx_desc.get('desc2', 2)] if 'desc2' in idx_desc else 0
            desc3 = row[idx_desc.get('desc3', 3)] if 'desc3' in idx_desc else 0
            desc4 = row[idx_desc.get('desc4', 4)] if 'desc4' in idx_desc else 0
            
            # Convertir a porcentaje (si están en decimal)
            descuentos_data[sku_str] = {
                'desc1': float(desc1) if desc1 else 0,
                'desc2': float(desc2) if desc2 else 0,
                'desc3': float(desc3) if desc3 else 0,
                'desc4': float(desc4) if desc4 else 0,
            }
        
        print(f"Descuentos leídos: {len(descuentos_data)} (filtrados: {descuentos_filtrados}, ignorados: {descuentos_ignorados})")
        
        # =====================================================
        # Leer hoja PRECIOS_FIJOS (remates/liquidaciones)
        # =====================================================
        ws_fijos = wb['PRECIOS_FIJOS']
        precios_fijos_data = {}
        
        # Obtener encabezados
        headers_fijos = [str(c.value).strip().lower() if c.value else "" for c in ws_fijos[1]]
        idx_fijos = {h: i for i, h in enumerate(headers_fijos) if h}
        
        # Verificar columnas requeridas
        if 'sku' not in idx_fijos:
            errs.append("Falta columna 'sku' en hoja PRECIOS_FIJOS")
            return precios, errs
        
        if 'precio_fijo' not in idx_fijos:
            errs.append("Falta columna 'precio_fijo' en hoja PRECIOS_FIJOS")
            return precios, errs
        
        # Leer precios fijos (solo los que están en catálogo)
        precios_fijos_filtrados = 0
        precios_fijos_ignorados = 0
        
        for row in ws_fijos.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            
            sku = row[idx_fijos.get('sku', 0)]
            precio_fijo = row[idx_fijos.get('precio_fijo', 1)]
            
            if not sku:
                continue
            
            sku_str = str(sku).strip()
            
            # FILTRO: Solo incluir precios fijos de SKUs que están en PRODUCTOS
            if sku_str not in productos_data:
                precios_fijos_ignorados += 1
                continue
            
            precios_fijos_filtrados += 1
            
            if precio_fijo:
                precios_fijos_data[sku_str] = float(precio_fijo)
        
        print(f"Precios fijos leídos: {len(precios_fijos_data)} (filtrados: {precios_fijos_filtrados}, ignorados: {precios_fijos_ignorados})")
        
        # =====================================================
        # Combinar datos
        # =====================================================
        for sku, orden in productos_data.items():
            p = {
                'orden': orden,
                'sku': sku,
            }
            
            # Agregar sku_cliente si existe
            if sku in sku_clientes_data:
                p['sku_cliente'] = sku_clientes_data[sku]
            
            # Si tiene precio fijo, usarlo (REMATe/liquidación)
            if sku in precios_fijos_data:
                p['precio_fijo'] = precios_fijos_data[sku]
                p['tipo'] = 'FIJO'
                p['es_remate'] = True  # Marcar como remate/liquidación
            else:
                p['precio_fijo'] = None
                p['tipo'] = 'DESCUENTO'
                p['es_remate'] = False
            
            # Agregar descuentos si existen
            if sku in descuentos_data:
                desc = descuentos_data[sku]
                p['desc1'] = desc['desc1']
                p['desc2'] = desc['desc2']
                p['desc3'] = desc['desc3']
                p['desc4'] = desc['desc4']
            else:
                p['desc1'] = 0
                p['desc2'] = 0
                p['desc3'] = 0
                p['desc4'] = 0
            
            precios.append(p)
        
        wb.close()
    except Exception as e:
        errs.append(str(e))
    return precios, errs


def gen_meta(precios, ruta):
    tipos = sorted(set(p.get("tipo", "DESCUENTO") for p in precios))
    total_remates = sum(1 for p in precios if p.get('es_remate', False))
    return {
        "version": "1.1.0",
        "generated_at": datetime.now().isoformat(),
        "source_file": Path(ruta).name,
        "total_precios": len(precios),
        "tipos_precio": tipos,
        "moneda_default": "PEN",
        "formato": "plantilla_precios",
        "total_remates": total_remates,
        "nota": "Productos con 'es_remate': true son remates/liquidaciones y deben resaltarse en reportes"
    }


def guardar(datos, ruta):
    try:
        p = Path(ruta)
        if p.exists():
            bk = p.with_suffix(f".{datetime.now():%Y%m%d_%H%M%S}.bak.json")
            p.rename(bk)
            print(f"Backup: {bk.name}")
        with open(p, "w", encoding="utf-8") as f:
            json.dump(datos, f, ensure_ascii=False, indent=2)
        print(f"Guardado: {ruta}")
        return True
    except Exception as e:
        print(f"Error: {e}")
        return False


def main():
    ap = argparse.ArgumentParser(description="G360 Plantilla Precios Reader v1.0")
    ap.add_argument("--input", "-i", required=True, help="Excel plantilla_precios.xlsx")
    ap.add_argument("--output", "-o", default="precios_productos.json", help="JSON salida")
    ap.add_argument("--validate-only", "-v", action="store_true")
    
    a = ap.parse_args()
    
    print("=" * 50)
    print("G360 PLANTILLA PRECIOS READER v1.0")
    print("=" * 50)
    print(f"Input:  {a.input}")
    print(f"Output: {a.output}")
    print()
    
    precios, errs = leer_plantilla_precios(a.input)
    
    if errs:
        print(f"\nErrores ({len(errs)}):")
        for e in errs[:10]:
            print(f"  - {e}")
        if len(errs) > 10:
            print(f"  ...y {len(errs) - 10} mas")
    
    if not precios:
        print("\nSin precios validos")
        sys.exit(1)
    
    print(f"\nOK: {len(precios)} precios")
    print(f"Con descuentos: {sum(1 for p in precios if p['tipo'] == 'DESCUENTO')}")
    print(f"Precios fijos: {sum(1 for p in precios if p['tipo'] == 'FIJO')}")
    
    if a.validate_only:
        print("\nValidacion completada")
        return
    
    meta = gen_meta(precios, a.input)
    if guardar({"metadata": meta, "precios": precios}, a.output):
        print(f"\n{'=' * 50}")
        print("PRECIOS GENERADOS")
        print(f"{'=' * 50}")
        print(f"Total: {meta['total_precios']}")
        print(f"Tipos: {', '.join(meta['tipos_precio'])}")


if __name__ == "__main__":
    main()