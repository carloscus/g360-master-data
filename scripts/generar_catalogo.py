#!/usr/bin/env python3
"""
G360 Catalog Generator
Genera JSON desde Excel para proyectos G360
Autor: CCUSI | v1.0.0
"""
import argparse
import json
import sys
import re
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: pip install openpyxl")
    sys.exit(1)

COLUMNS = [
    "orden", "sku", "nombre", "ean13", "categoria", "estado_linea", "subcategoria",
    "linea", "grupo", "tipo", "familia", "unidad_medida", "peso_kg",
    "un_bx", "precio_lista", "moneda", "imagen_url"
]
DEFAULTS = {"moneda":"PEN","unidad_medida":"UND"}

# Mapeo de línea/categoría a grupo
GRUPO_MAP = {
    # Por categoría (marca)
    'VINIBALL': 'DEPORTES',
    'VINIFAN': 'ESCOLAR',
    'REPRESENTADAS': 'REPRESENTADAS',
    # Por línea (compatibilidad)
    'PELOTAS': 'DEPORTES',
    'MASCOTAS': 'DEPORTES',
    'FORROS': 'ESCOLAR',
    'ARCHIVO': 'ESCOLAR',
    'PINTURA': 'ESCOLAR',
    'ESCRITURA': 'ESCOLAR',
    'ACCESORIOS': 'ESCOLAR',
    'DIBUJO': 'ESCOLAR',
    'DIDACTICOS': 'ESCOLAR',
    'MANUALIDADES': 'ESCOLAR',
    'PEGAMENTOS': 'ESCOLAR',
    'ESCOLAR': 'ESCOLAR',
}

def generar_keywords(nombre, linea, categoria, subcategoria):
    """Genera keywords automáticamente desde nombre y categorías"""
    keywords = set()
    
    # Agregar palabras del nombre
    if nombre:
        palabras = nombre.upper().replace("#","").replace("."," ").split()
        for p in palabras:
            if len(p) > 2:  # Ignorar palabras muy cortas
                keywords.add(p)
    
    # Agregar línea y categoría
    if linea:
        keywords.add(linea.upper())
    if categoria:
        keywords.add(categoria.upper())
    if subcategoria:
        keywords.add(subcategoria.upper())
    
    return sorted(list(keywords))

def generar_nombre_corto(nombre):
    """
    Genera un nombre corto a partir de un nombre largo.
    
    Reglas:
    1. Eliminar 'PELOTA' o 'PELOTA DE' al inicio
    2. Eliminar materiales: GOMA, PVC, CUERO, PU
    3. Eliminar marcas: VINIFAN*, VFAN*
    4. Eliminar 'N' sola (en cualquier posición)
    5. Eliminar espacios múltiples
    6. Capitalizar
    """
    if not nombre:
        return ""
    
    # Convertir a mayúsculas para procesamiento
    texto = nombre.upper().strip()
    
    # 1. Eliminar 'PELOTA DE' o 'PELOTA' al inicio
    texto = re.sub(r'^PELOTA(\s+DE)?\s+', '', texto)
    
    # 2. Eliminar materiales
    texto = re.sub(r'\b(GOMA|PVC|CUERO|PU)\b', '', texto)
    
    # 3. Eliminar marcas (VINIFAN* y VFAN*)
    texto = re.sub(r'\bVINIFAN\w*\b', '', texto)
    texto = re.sub(r'\bVFAN\w*\b', '', texto)
    
    # 4. Eliminar 'N' sola (palabra completa)
    texto = re.sub(r'\bN\b', '', texto)
    
    # 5. Limpiar espacios múltiples
    texto = re.sub(r'\s+', ' ', texto).strip()
    
    # 6. Capitalizar primera letra de cada palabra
    texto = texto.title()
    
    return texto

def validar_ean13(ean):
    if not ean or ean=="0": return True
    ean=str(ean).strip()
    if len(ean)!=13 or not ean.isdigit(): return False
    t=sum(int(d)*(1 if i%2==0 else 3) for i,d in enumerate(ean[:12]))
    return (10-(t%10))%10==int(ean[12])

def leer_excel(ruta):
    """
    Lee el archivo Excel y genera la lista de productos consolidados.
    
    Validaciones:
    - SKU y nombre obligatorios (si no, se salta)
    - categoria y linea obligatorios (si no, se salta)
    - SKUs duplicados: solo se considera el primero
    - un_bx vacío: se asigna 1
    - precio_lista vacío: se asigna 0.00
    - ean13 vacío: se incluye sin EAN (observación)
    
    Returns:
        prods: Lista de productos procesados
        errs: Lista de errores encontrados
    """
    prods = []
    errs = []
    observaciones = []
    descuentos_map = {}
    precios_fijos_map = {}
    skus_vistos = set()  # Para detectar duplicados
    
    try:
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        
        # =====================================================
        # Leer hoja DESCUENTOS (dinámico: soporta N columnas)
        # =====================================================
        if 'DESCUENTOS' in wb.sheetnames:
            ws_desc = wb['DESCUENTOS']
            hdrs_desc = [str(c.value).strip().lower() if c.value else "" for c in ws_desc[1]]
            idx_desc = {h: i for i, h in enumerate(hdrs_desc) if h}
            
            # Detectar columnas de descuento dinámicamente (desc1, desc2, desc3, etc.)
            desc_cols = sorted([h for h in idx_desc.keys() if h.startswith('desc') and h[4:].isdigit()])
            
            for row in ws_desc.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                sku = row[idx_desc.get('sku', 0)]
                if not sku:
                    continue
                sku = str(sku).strip()
                
                # Leer descuentos dinámicamente
                descuentos = []
                for col in desc_cols:
                    val = row[idx_desc.get(col, 0)] if idx_desc.get(col, 0) < len(row) else 0
                    descuentos.append(float(val or 0))
                
                # Si no hay columnas desc, usar [0,0,0,0] por defecto
                if not descuentos:
                    descuentos = [0, 0, 0, 0]
                
                descuentos_map[sku] = descuentos
        
        # =====================================================
        # Leer hoja PRECIOS_FIJOS
        # =====================================================
        if 'PRECIOS_FIJOS' in wb.sheetnames:
            ws_fijos = wb['PRECIOS_FIJOS']
            hdrs_fijos = [str(c.value).strip().lower() if c.value else "" for c in ws_fijos[1]]
            idx_fijos = {h: i for i, h in enumerate(hdrs_fijos) if h}
            
            for row in ws_fijos.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                sku = row[idx_fijos.get('sku', 0)]
                precio_fijo = row[idx_fijos.get('precio_fijo', 1)]
                if not sku:
                    continue
                sku = str(sku).strip()
                if precio_fijo:
                    precios_fijos_map[sku] = float(precio_fijo)
        
        # =====================================================
        # Leer hoja PRODUCTOS (principal)
        # =====================================================
        ws = wb.active
        hdrs = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        
        # Verificar columnas obligatorias
        for r in ["orden", "sku", "nombre", "linea", "categoria"]:
            if r not in hdrs:
                errs.append(f"Falta columna obligatoria: {r}")
        if errs:
            return prods, errs
        
        idx = {h: i for i, h in enumerate(hdrs) if h in COLUMNS}
        
        for rn, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(row):
                continue
            
            p = {}
            for c in COLUMNS:
                vi = idx.get(c)
                v = row[vi] if vi is not None and vi < len(row) else None
                if isinstance(v, str):
                    v = v.strip()
                
                # Procesar según tipo de campo
                if c in ["peso_kg", "precio_lista"]:
                    p[c] = round(float(v), 2) if v not in (None, "") else 0.0
                elif c in ["un_bx", "orden"]:
                    # un_bx vacío se asigna 1
                    if c == "un_bx" and v in (None, ""):
                        p[c] = 1
                    else:
                        p[c] = int(v) if v not in (None, "") else 0
                else:
                    p[c] = v or ""
            
            # =====================================================
            # Validaciones obligatorias
            # =====================================================
            sku_val = p.get("sku", "").strip()
            nombre_val = p.get("nombre", "").strip()
            categoria_val = p.get("categoria", "").strip()
            linea_val = p.get("linea", "").strip()
            
            # SKU vacío: saltar
            if not sku_val:
                errs.append(f"Fila {rn}: SKU vacío - producto ignorado")
                continue
            
            # Nombre vacío: saltar
            if not nombre_val:
                errs.append(f"Fila {rn}: Nombre vacío - producto ignorado")
                continue
            
            # Categoria vacía: saltar
            if not categoria_val:
                errs.append(f"Fila {rn}: Categoria vacía - producto ignorado")
                continue
            
            # Linea vacía: saltar
            if not linea_val:
                errs.append(f"Fila {rn}: Linea vacía - producto ignorado")
                continue
            
            # SKU duplicado: saltar (solo se considera el primero)
            if sku_val in skus_vistos:
                observaciones.append(f"SKU duplicado ignorado: {sku_val}")
                continue
            skus_vistos.add(sku_val)
            
            # EAN vacío: observación (no error)
            ean_val = p.get("ean13", "")
            if not ean_val or ean_val == "0":
                observaciones.append(f"SKU {sku_val}: sin EAN13")
            elif not validar_ean13(ean_val):
                errs.append(f"Fila {rn}: EAN13 inválido")
            
            # Valores por defecto
            p.setdefault("moneda", DEFAULTS["moneda"])
            p.setdefault("unidad_medida", DEFAULTS["unidad_medida"])
            
            # Generar keywords automáticamente
            p["keywords"] = generar_keywords(
                p.get("nombre", ""),
                p.get("linea", ""),
                p.get("categoria", ""),
                p.get("subcategoria", "")
            )
            
            # Generar nombre_corto
            p["nombre_corto"] = generar_nombre_corto(p.get("nombre", ""))
            
            # Agregar descuentos
            sku = p.get("sku", "")
            descuentos = descuentos_map.get(sku, [0, 0, 0, 0])
            p["descuentos"] = descuentos
            
            # Calcular descuento total (consecutivo, no acumulativo)
            factor = 1
            for d in descuentos:
                if d > 0:
                    factor *= (1 - d)
            desc_total = round((1 - factor) * 100, 1) if factor < 1 else 0
            p["descuento_pct"] = desc_total
            
            # Agregar precio_fijo si existe
            precio_fijo = precios_fijos_map.get(sku)
            p["precio_fijo"] = precio_fijo
            
            # Calcular precio_final
            # Prioridad: precio_fijo > descuentos > precio_lista
            if precio_fijo:
                p["precio_final"] = precio_fijo
                p["es_remate"] = True
            elif desc_total > 0:
                p["precio_final"] = round(p["precio_lista"] * (1 - desc_total / 100), 2)
                p["es_remate"] = False
            else:
                p["precio_final"] = p["precio_lista"]
                p["es_remate"] = False
            
            prods.append(p)
        
        wb.close()
        
        # Mostrar observaciones
        if observaciones:
            print(f"\n⚠️  Observaciones ({len(observaciones)}):")
            for obs in observaciones[:10]:
                print(f"  - {obs}")
            if len(observaciones) > 10:
                print(f"  ...y {len(observaciones) - 10} más")
        
    except Exception as e:
        errs.append(str(e))
    
    return prods, errs

def gen_meta(prods,ruta):
    """Genera metadata del catálogo con versionado y changelog"""
    lns=sorted(set(p.get("linea","") for p in prods if p.get("linea")))
    cats=sorted(set(p.get("categoria","") for p in prods if p.get("categoria")))
    mons=[p.get("moneda","PEN") for p in prods]
    
    # Contar observaciones
    sin_ean=sum(1 for p in prods if not p.get("ean13") or p.get("ean13")=="0")
    con_descuento=sum(1 for p in prods if p.get("descuento_pct",0)>0)
    con_precio_fijo=sum(1 for p in prods if p.get("es_remate",False))
    
    return {
        "version": "2.0.0",
        "generated_at": datetime.now().isoformat(),
        "source_file": Path(ruta).name,
        "total_productos": len(prods),
        "lineas": lns,
        "categorias": cats,
        "moneda_default": max(set(mons), key=mons.count),
        "estadisticas": {
            "sin_ean13": sin_ean,
            "con_descuento": con_descuento,
            "con_precio_fijo": con_precio_fijo
        },
        "changelog": [
            "v2.0.0: Descuentos dinámicos (N columnas)",
            "v2.0.0: Validaciones mejoradas (SKU/nombre/categoria/linea obligatorios)",
            "v2.0.0: SKUs duplicados ignorados",
            "v2.0.0: un_bx vacío asigna 1",
            "v2.0.0: nombre_corto generado con regex",
            "v2.0.0: Consolidación de descuentos y precios fijos"
        ]
    }

def guardar(datos,ruta):
    try:
        p=Path(ruta)
        if p.exists():
            bk=p.with_suffix(f".{datetime.now():%Y%m%d_%H%M%S}.bak.json")
            p.rename(bk); print(f"Backup: {bk.name}")
        with open(p,"w",encoding="utf-8") as f: json.dump(datos,f,ensure_ascii=False,indent=2)
        print(f"Guardado: {ruta}"); return True
    except Exception as e: print(f"Error: {e}"); return False

def gen_template(ruta="plantilla_catalogo.xlsx"):
    try:
        wb=openpyxl.Workbook()
        wb.properties.creator="ccusi"
        wb.properties.description="Generado por G360"
        ws=wb.active; ws.title="Catalogo"
        hdrs=[("orden",8),("sku",15),("nombre",40),("linea",20),("categoria",20),
              ("subcategoria",20),("ean13",18),("unidad_medida",12),("peso_kg",10),
              ("un_bx",10),("precio_lista",12),("moneda",8),("imagen_url",30)]
        for c,(h,w) in enumerate(hdrs,1):
            ws.cell(1,c,h).font=openpyxl.styles.Font(bold=True)
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width=w
        ej=[1,"016763","FUTBOL PU FUTURE #5","PELOTAS","DEPORTES","FUTBOL",
            "7754807167277","UND",0.42,18,25.50,"PEN",""]
        for c,v in enumerate(ej,1): ws.cell(2,c,v)
        wb.save(ruta); print(f"Plantilla: {ruta}")
    except Exception as e: print(f"Error: {e}")

def main():
    ap=argparse.ArgumentParser(description="G360 Catalog Generator")
    ap.add_argument("--input","-i",help="Excel entrada")
    ap.add_argument("--output","-o",default="catalogo_productos.json",help="JSON salida")
    ap.add_argument("--validate-only","-v",action="store_true")
    ap.add_argument("--generate-template","-t",action="store_true")
    a=ap.parse_args()
    if a.generate_template:
        print("="*50+"\nGENERANDO PLANTILLA\n"+"="*50); gen_template(); return
    if not a.input: ap.error("Requiere --input")
    print("="*50+"\nG360 CATALOG GENERATOR\n"+"="*50)
    print(f"Input:  {a.input}\nOutput: {a.output}\n")
    prods,errs=leer_excel(a.input)
    if errs:
        print(f"Errores ({len(errs)}):")
        for e in errs[:10]: print(f"  - {e}")
        if len(errs)>10: print(f"  ...y {len(errs)-10} mas")
    if not prods: print("Sin productos validos"); sys.exit(1)
    print(f"OK: {len(prods)} productos")
    if a.validate_only: print("Validacion completada"); return
    meta=gen_meta(prods,a.input)
    if guardar({"metadata":meta,"productos":prods},a.output):
        print(f"\n{'='*50}\nCATALOGO GENERADO\n{'='*50}")
        print(f"Productos: {meta['total_productos']}")
        print(f"Lineas: {', '.join(meta['lineas'][:5])}")

if __name__=="__main__": main()